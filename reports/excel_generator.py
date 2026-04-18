"""Daily report generator — produces a multi-sheet Excel workbook."""

import json
import os
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────

_GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
_GREEN_FONT  = Font(color="276221", bold=True)
_YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
_YELLOW_FONT = Font(color="9C5700", bold=True)
_RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
_RED_FONT    = Font(color="9C0006", bold=True)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

_SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
_SECTION_FONT = Font(bold=True, color="1F4E79", size=11)

_LABEL_FONT  = Font(bold=True)
_WRAP        = Alignment(wrap_text=True, vertical="top")
_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)

_SCORE_PARAMS = [
    ("news_sentiment",      "News"),
    ("buffett_value",       "Buffett"),
    ("geopolitical_thesis", "Thesis"),
    ("self_critique",       "Critique"),
    ("govt_contracts",      "Contracts"),
    ("pnl_trend",           "P&L"),
    ("politician_trades",   "Politicians"),
    ("competitor_analysis", "Competitors"),
]


# ── Style helpers ─────────────────────────────────────────────────────────────

def _score_style(score):
    """Return (fill, font) based on numeric score."""
    try:
        s = float(score)
    except (TypeError, ValueError):
        return None, None
    if s >= 7:
        return _GREEN_FILL, _GREEN_FONT
    if s >= 4:
        return _YELLOW_FILL, _YELLOW_FONT
    return _RED_FILL, _RED_FONT


def _apply_score(cell, score):
    fill, font = _score_style(score)
    if fill:
        cell.fill = fill
        cell.font = font


def _header_row(ws, row_idx: int):
    for cell in ws[row_idx]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER


def _section_row(ws, row_idx: int):
    for cell in ws[row_idx]:
        if cell.value:
            cell.fill = _SECTION_FILL
            cell.font = _SECTION_FONT
            cell.alignment = Alignment(vertical="center")


def _autofit(ws, max_width: int = 80):
    for col in ws.columns:
        best = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(best + 4, max_width)


def _set_row_height(ws, row_idx: int, height: float):
    ws.row_dimensions[row_idx].height = height


# ── Sheet 1: Summary ──────────────────────────────────────────────────────────

def _build_summary(wb: Workbook, picks: list, report_date: str):
    ws = wb.active
    ws.title = "Summary"

    # Title row
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = f"Stock Sentinel — Daily Report — {report_date}"
    title_cell.font  = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = _CENTER
    _set_row_height(ws, 1, 24)

    # Header row
    headers = [
        "Rank", "Ticker", "Company", "Recommendation", "Composite",
        "News", "Buffett", "Thesis", "Critique",
        "Contracts", "P&L", "Politicians", "Competitors",
        "One-Line Summary",
    ]
    ws.append(headers)
    _header_row(ws, 2)
    _set_row_height(ws, 2, 18)

    for rank, a in enumerate(picks, start=1):
        scores = a.get("scores", {})

        def _s(key):
            return scores.get(key, {}).get("score", "")

        rec = a.get("recommendation", "")
        composite = a.get("composite_score", "")

        row_vals = [
            rank,
            a.get("ticker", ""),
            a.get("company_name", ""),
            rec,
            composite,
            _s("news_sentiment"),
            _s("buffett_value"),
            _s("geopolitical_thesis"),
            _s("self_critique"),
            _s("govt_contracts"),
            _s("pnl_trend"),
            _s("politician_trades"),
            _s("competitor_analysis"),
            a.get("one_line_summary", ""),
        ]
        ws.append(row_vals)
        data_row = rank + 2  # offset: title + header

        # Colour composite and all score columns (cols 5-13)
        _apply_score(ws.cell(data_row, 5), composite)
        for col_offset, (param_key, _) in enumerate(_SCORE_PARAMS, start=6):
            _apply_score(ws.cell(data_row, col_offset), _s(param_key))

        # Colour recommendation cell
        rec_cell = ws.cell(data_row, 4)
        rec_cell.font = Font(bold=True)
        if rec == "Buy":
            rec_cell.fill = _GREEN_FILL; rec_cell.font = _GREEN_FONT
        elif rec == "Hold":
            rec_cell.fill = _YELLOW_FILL; rec_cell.font = _YELLOW_FONT
        elif rec == "Avoid":
            rec_cell.fill = _RED_FILL; rec_cell.font = _RED_FONT

        ws.cell(data_row, 14).alignment = _WRAP
        _set_row_height(ws, data_row, 40)

    ws.freeze_panes = "A3"
    _autofit(ws)
    # Override wide columns
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["N"].width = 60


# ── Sheet per stock ───────────────────────────────────────────────────────────

def _blank(ws, n: int = 1):
    for _ in range(n):
        ws.append([""])


def _section(ws, title: str):
    ws.append([title])
    _section_row(ws, ws.max_row)
    _set_row_height(ws, ws.max_row, 20)


def _kv(ws, label: str, value, wide: bool = True):
    ws.append([label, value])
    r = ws.max_row
    ws.cell(r, 1).font = _LABEL_FONT
    ws.cell(r, 2).alignment = _WRAP
    if wide:
        _set_row_height(ws, r, max(15, min(len(str(value or "")) // 5, 80)))


def _fmt_large(v) -> str:
    if v is None:
        return "N/A"
    try:
        f = float(v)
        if abs(f) >= 1e12: return f"${f/1e12:.2f}T"
        if abs(f) >= 1e9:  return f"${f/1e9:.2f}B"
        if abs(f) >= 1e6:  return f"${f/1e6:.2f}M"
        return f"{f:,.2f}"
    except (TypeError, ValueError):
        return str(v)


def _fmt(v, prefix="", suffix="", decimals=2) -> str:
    if v is None:
        return "N/A"
    try:
        return f"{prefix}{float(v):.{decimals}f}{suffix}"
    except (TypeError, ValueError):
        return str(v)


def _build_stock_sheet(wb: Workbook, a: dict, rank: int):
    ticker = a.get("ticker", "UNKNOWN")
    ws = wb.create_sheet(ticker)
    scores = a.get("scores", {})
    fd = a.get("fundamentals", {})

    # ── Company overview ──────────────────────────────────────────────────────
    _section(ws, "COMPANY OVERVIEW")
    _kv(ws, "Ticker",          ticker,                          wide=False)
    _kv(ws, "Company",         a.get("company_name", ""),       wide=False)
    _kv(ws, "Sector",          fd.get("sector", "N/A"),         wide=False)
    _kv(ws, "Industry",        fd.get("industry", "N/A"),       wide=False)
    _kv(ws, "Recommendation",  a.get("recommendation", ""),     wide=False)
    _kv(ws, "Composite Score", a.get("composite_score", ""),    wide=False)
    _kv(ws, "Rank",            rank,                            wide=False)
    _blank(ws)
    _kv(ws, "Summary", a.get("one_line_summary", ""))
    _blank(ws)

    # ── Score breakdown ───────────────────────────────────────────────────────
    _section(ws, "SCORE BREAKDOWN")
    ws.append(["Parameter", "Score", "Key Reasoning"])
    _header_row(ws, ws.max_row)

    reasoning_keys = {
        "news_sentiment":      "reasoning",
        "buffett_value":       "reasoning",
        "geopolitical_thesis": "thesis",
        "self_critique":       "critique",
        "govt_contracts":      "reasoning",
        "pnl_trend":           "reasoning",
        "politician_trades":   "reasoning",
        "competitor_analysis": "reasoning",
    }
    for param_key, label in _SCORE_PARAMS:
        sub = scores.get(param_key, {})
        score = sub.get("score", "")
        reason_key = reasoning_keys.get(param_key, "reasoning")
        reasoning = sub.get(reason_key, "")
        ws.append([label, score, reasoning])
        r = ws.max_row
        _apply_score(ws.cell(r, 2), score)
        ws.cell(r, 3).alignment = _WRAP
        _set_row_height(ws, r, max(30, min(len(str(reasoning)) // 4, 90)))
    _blank(ws)

    # ── Buffett value detail ──────────────────────────────────────────────────
    _section(ws, "BUFFETT VALUE ANALYSIS")
    bv = scores.get("buffett_value", {})
    metrics = [
        ("Current Price",     _fmt(fd.get("current_price"), "$")),
        ("Market Cap",        _fmt_large(fd.get("market_cap"))),
        ("P/E Ratio",         _fmt(fd.get("pe_ratio"))),
        ("P/B Ratio",         _fmt(fd.get("pb_ratio"))),
        ("Free Cash Flow",    _fmt_large(fd.get("free_cash_flow"))),
        ("Debt / Equity",     _fmt(fd.get("debt_to_equity"))),
        ("Revenue TTM",       _fmt_large(fd.get("revenue_ttm"))),
        ("Net Income TTM",    _fmt_large(fd.get("net_income_ttm"))),
        ("EPS Growth YoY",    _fmt(fd.get("earnings_growth_yoy"), suffix="%")),
        ("52w High",          _fmt(fd.get("week_52_high"), "$")),
        ("52w Low",           _fmt(fd.get("week_52_low"),  "$")),
        ("Analyst Consensus", str(fd.get("analyst_consensus") or "N/A").replace("_", " ").title()),
        ("30d Price Change",  _fmt(fd.get("price_change_30d_pct"), suffix="%")),
    ]
    for label, val in metrics:
        _kv(ws, label, val, wide=False)

    _blank(ws)
    ws.append(["Assessment", "Detail"])
    _header_row(ws, ws.max_row)
    for key, label in [
        ("pe_assessment",   "P/E Assessment"),
        ("pb_assessment",   "P/B Assessment"),
        ("fcf_assessment",  "FCF Assessment"),
        ("debt_assessment", "Debt Assessment"),
    ]:
        text = bv.get(key, "")
        ws.append([label, text])
        ws.cell(ws.max_row, 2).alignment = _WRAP
        _set_row_height(ws, ws.max_row, max(25, min(len(str(text)) // 4, 80)))

    # Quarterly revenue table
    rev_q = fd.get("revenue_quarterly", [])
    if rev_q:
        _blank(ws)
        ws.append(["Quarterly Revenue"])
        _header_row(ws, ws.max_row)
        for q in rev_q:
            ws.append([q.get("period", ""), _fmt_large(q.get("value"))])

    ni_q = fd.get("net_income_quarterly", [])
    if ni_q:
        _blank(ws)
        ws.append(["Quarterly Net Income"])
        _header_row(ws, ws.max_row)
        for q in ni_q:
            ws.append([q.get("period", ""), _fmt_large(q.get("value"))])

    _blank(ws)

    # ── Investment thesis ─────────────────────────────────────────────────────
    _section(ws, "INVESTMENT THESIS (GEOPOLITICAL)")
    gt = scores.get("geopolitical_thesis", {})
    _kv(ws, "Score",  gt.get("score", ""), wide=False)
    _kv(ws, "Thesis", gt.get("thesis", ""))
    _blank(ws)

    pnl = scores.get("pnl_trend", {})
    _section(ws, "P&L TREND")
    _kv(ws, "Score",          pnl.get("score", ""),          wide=False)
    _kv(ws, "Revenue Trend",  pnl.get("revenue_trend", ""),  wide=False)
    _kv(ws, "Margin Trend",   pnl.get("margin_trend", ""),   wide=False)
    _kv(ws, "Reasoning",      pnl.get("reasoning", ""))
    _blank(ws)

    # ── Self-critique ─────────────────────────────────────────────────────────
    _section(ws, "COUNTER-THESIS / SELF-CRITIQUE")
    sc = scores.get("self_critique", {})
    _kv(ws, "Score",    sc.get("score", ""),   wide=False)
    _kv(ws, "Critique", sc.get("critique", ""))
    _blank(ws)

    # ── Politician trades ─────────────────────────────────────────────────────
    _section(ws, "POLITICIAN TRADE NOTES")
    pt = scores.get("politician_trades", {})
    _kv(ws, "Score",     pt.get("score", ""),     wide=False)
    _kv(ws, "Reasoning", pt.get("reasoning", ""))
    _blank(ws)

    # ── Competitor landscape ──────────────────────────────────────────────────
    _section(ws, "COMPETITOR LANDSCAPE")
    ca = scores.get("competitor_analysis", {})
    _kv(ws, "Score",            ca.get("score", ""),     wide=False)
    competitors = ca.get("main_competitors", [])
    _kv(ws, "Main Competitors", ", ".join(competitors) if competitors else "N/A", wide=False)
    _kv(ws, "Reasoning",        ca.get("reasoning", ""))
    _blank(ws)

    # ── Govt contracts ────────────────────────────────────────────────────────
    _section(ws, "GOVERNMENT CONTRACTS")
    gc = scores.get("govt_contracts", {})
    _kv(ws, "Score",     gc.get("score", ""),     wide=False)
    _kv(ws, "Reasoning", gc.get("reasoning", ""))
    _blank(ws)

    # ── Reddit context ────────────────────────────────────────────────────────
    _section(ws, "REDDIT MENTION CONTEXT")
    reddit = a.get("reddit_context", {})
    _kv(ws, "24h Mentions",   reddit.get("mentions", "N/A"),        wide=False)
    _kv(ws, "Subreddits",     reddit.get("subreddits_seen", "N/A"), wide=False)
    top_posts = reddit.get("top_posts", [])
    if top_posts:
        ws.append(["Top Post Titles", ""])
        _header_row(ws, ws.max_row)
        for post in top_posts[:3]:
            ws.append(["", post])
            ws.cell(ws.max_row, 2).alignment = _WRAP

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 70


# ── Public entry point ────────────────────────────────────────────────────────

def generate_report(
    analyses: list,
    reddit_rows: Optional[list] = None,
    output_path=None,
    top_n: Optional[int] = None,
    report_date: Optional[str] = None,
) -> Path:
    """Build and save the daily Excel workbook. Returns the saved Path.

    Args:
        analyses:     list of dicts from analyse_stock() — must have 'composite_score'.
        reddit_rows:  raw reddit mention rows (optional) — merged in for context.
        output_path:  override destination; defaults to data/reports/YYYY-MM-DD_stock_sentinel.xlsx
        top_n:        max tickers to include (default: TOP_N_FINAL env var or 5).
        report_date:  ISO date string; defaults to today.
    """
    if top_n is None:
        top_n = int(os.getenv("TOP_N_FINAL", 5))
    if report_date is None:
        report_date = date.today().isoformat()

    if output_path is None:
        out_dir = Path(os.path.dirname(__file__)) / ".." / "data" / "reports"
        out_dir.mkdir(parents=True, exist_ok=True)
        output_path = out_dir / f"{report_date}_stock_sentinel.xlsx"
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Build reddit lookup for merging context into per-stock sheets
    reddit_map = {}
    if reddit_rows:
        for r in reddit_rows:
            t = r.get("ticker", "")
            if t:
                reddit_map[t] = r

    # Sort and select top N valid analyses
    valid = [a for a in analyses if not a.get("error")]
    picks = sorted(valid, key=lambda x: float(x.get("composite_score") or 0), reverse=True)[:top_n]

    # Attach reddit context to each pick so per-stock sheets can show it
    for pick in picks:
        t = pick.get("ticker", "")
        pick["reddit_context"] = reddit_map.get(t, pick.get("reddit_context", {}))

    wb = Workbook()
    _build_summary(wb, picks, report_date)

    for rank, a in enumerate(picks, start=1):
        _build_stock_sheet(wb, a, rank)

    wb.save(output_path)
    return output_path
